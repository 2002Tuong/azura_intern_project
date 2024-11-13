import os
import re

import openpyxl

def readExcelFile(filePath):
    dataFrame = openpyxl.load_workbook(filePath, data_only=True)
    frame = dataFrame["Call Theme Text"]
    for row in frame.iter_rows(2, frame.max_row):
        key = row[0].value
        if key:
            for index, col in enumerate(frame.iter_cols(1, frame.max_column)):
                lang = col[1].value
                if lang:
                    if lang == "en":
                        folder = "values"
                    else:
                        folder = f"values-{lang.lower()}"
                    if not os.path.exists(folder):
                        os.mkdir(folder)
                    text = []
                    if os.path.exists("{}/strings.xml".format(folder)):
                        with open("{}/strings.xml".format(folder), "r") as file:
                            text = file.read().splitlines(True)[1:-1]
                        for line in text:
                            x = re.search(r"\s<string name=\"{}.*\">(\"?.+\"?)<\/string>\s".format(key), line)
                            # remove old string with same key
                            if x:
                                text = list(filter(lambda a: a != x.group(0), text))

                    with open("{}/strings.xml".format(folder), "w") as file:
                        file.write("<resources>\n")
                        file.writelines(text)
                        value = row[index].value
                        if value:
                            string_res = "<string name=\"{}\">\"{}\"</string>".format(key, value)
                            file.writelines(f"\t{string_res}\n")
                        file.write("</resources>\n")


readExcelFile("translation.xlsx")

