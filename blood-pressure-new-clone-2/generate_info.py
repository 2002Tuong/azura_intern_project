import openpyxl
import os
import shutil

def readExcelFile(filePath):
    dataFrame = openpyxl.load_workbook(filePath)
    frame = dataFrame["HTML - Info card"]

    keys = []
    for row in frame.iter_rows(3, frame.max_row):
        key = row[0].value
        if key:
            keys.append(key)

    language_codes = []
    for col in frame.iter_cols(2, frame.max_column):
        lang = col[1].value
        if lang:
            language_codes.append(lang)

    fileIndex = 0
    folder = "infos"
    if os.path.exists(folder):
        shutil.rmtree(folder)
    os.mkdir(folder)
    for col in range(1, len(language_codes) + 1):
        language_code = language_codes[fileIndex]
        print(language_code)
        print(col)
        file = open("{}/info-{}.json".format(folder, language_code), "x")
        file.write("{")
        key_index = 0
        for row in frame.iter_rows(4, len(keys) + 3):
            value = row[col].value
            key = keys[key_index]
            file.write("\n")
            string_res = "    \"{}\": \"{}\",".format(key, value)
            if key_index == len(keys) - 1:
               string_res = "    \"{}\": \"{}\" ".format(key, value)
            file.write(string_res)
            key_index = key_index + 1
        file.write("\n")
        file.write("}")
        file.close()
        fileIndex = fileIndex + 1
readExcelFile("bp_app.xlsx")
