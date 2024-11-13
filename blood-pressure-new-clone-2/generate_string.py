import openpyxl
import os
import shutil

def readExcelFile(filePath):
    dataFrame = openpyxl.load_workbook(filePath)
    frame = dataFrame["Inapp text"]

    keys = []
    for row in frame.iter_rows(3, frame.max_row):
        key = row[0].value
        if key:
            keys.append(key)

    language_codes = []
    for col in frame.iter_cols(2, frame.max_column):
        lang = col[1].value
        if lang:
            language_codes.append(lang)

    fileIndex = 0
    for col in range(1, len(language_codes) + 1):
        folder = "app/src/main/res/values-{}".format(language_codes[fileIndex])
        if os.path.exists(folder):
            shutil.rmtree(folder)
        os.mkdir(folder)
        file = open("{}/strings.xml".format(folder), "x")
        file.write("<resources>")
        key_index = 0
        for row in frame.iter_rows(3, len(keys) + 2) :
            value = row[col].value
            key = keys[key_index]
            file.write("\n")
            string_res = "    <string name=\"{}\">\"{}\"</string>".format(key, value)
            file.write(string_res)
            key_index = key_index + 1
        file.write("\n")
        file.write("</resources>")
        file.close()
        fileIndex = fileIndex + 1
readExcelFile("bp_app.xlsx")
