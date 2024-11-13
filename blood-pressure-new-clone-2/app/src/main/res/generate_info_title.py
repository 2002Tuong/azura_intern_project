import html
import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import re

# Path to the Excel file
excel_file_path = 'translation.xlsx'

feature = "water"

output_dir = '.'

# Load the Excel file and the first sheet
workbook = load_workbook(excel_file_path)
sheet = workbook.worksheets[0]  # Use the first sheet

# Get a list of available locales (column names)
locales = sheet[2]  # Assuming the locales are in the second row

# Skip the first column ('key') and iterate through locales
for col_idx in range(1, len(locales)):
    locale = locales[col_idx].value
    if locale is None:
        continue
    locale_dir = os.path.join(output_dir, f"values-{locale.lower()}")

    # Create the locale directory if it doesn't exist
    os.makedirs(locale_dir, exist_ok=True)

    # Create the root element for the XML file
    root = ET.Element('resources')

    # Iterate through rows and populate the XML content
    for row_idx in range(3, sheet.max_row + 1):
        key = sheet.cell(row=row_idx, column=1).value
        value = sheet.cell(row=row_idx, column=col_idx+1).value

        # Create a <string> element for each key-value pair
        h1_pattern = r"<h1>(.*?)</h1>"
        if key and value:
            match = re.search(h1_pattern, value, re.DOTALL)

            if match:
                h1_content = match.group(1)
                string_elem = ET.SubElement(root, 'string', name=f"{feature}_info_title_{key}")
                string_elem.text = html.unescape(h1_content)


    # Create the XML file for the current locale
    xml_tree = ET.ElementTree(root)
    xml_file_path = os.path.join(locale_dir, f'{feature}_title_info.xml')
    xml_tree.write(xml_file_path, encoding='utf-8', xml_declaration=True)

    print(f'Generated XML for {locale} in {xml_file_path}')